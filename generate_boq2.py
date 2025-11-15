#!/usr/bin/env python3
"""
Script to generate a brand new BOQ2 workbook with better tabularization
considering all special cases and changes made
"""

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# Special DBs and their copy count logic (matching script.js)
def get_db_copy_counts(parent_name):
    """Get copy counts for special DBs based on parent"""
    normalized = parent_name.replace('.01', '')
    
    # Special case for SMDB.TN.P2
    if normalized == 'SMDB.TN.P2':
        return {
            'DB.TN.LXX.1B1.01': 3,
            'DB.TN.LXX.2B1.01': 1,
            'DB.TN.LXX.3B1.01': 1
        }
    
    # Special case for DB.TH.GF.01 - appears under SMDB.TH.B1.01 and MDB4
    if normalized == 'SMDB.TH.B1' or parent_name == 'SMDB.TH.B1.01':
        return {
            'DB.TH.GF.01': 1
        }
    
    # Handle MDB4 (MDB.GF.04) as parent for DB.TH.GF.01
    if normalized == 'MDB.GF.04' or normalized == 'MDB4' or parent_name in ['MDB4', 'MDB.GF.04']:
        return {
            'DB.TH.GF.01': 1
        }
    
    # Handle SMDB.TN.L## pattern
    import re
    match = re.match(r'^SMDB\.TN\.L(\d+)$', normalized)
    if not match:
        return None
    
    level = int(match.group(1))
    
    if level == 1:
        return {
            'DB.TN.LXX.1B1.01': 5,
            'DB.TN.LXX.2B1.01': 3,
            'DB.TN.LXX.3B1.01': 1
        }
    
    if level == 23:
        return {
            'DB.TN.LXX.2B1.01': 2,
            'DB.TN.LXX.1B1.01': 1,
            'DB.TN.LXX.3B1.01': 1
        }
    
    if (level >= 2 and level <= 22) or (level >= 24 and level <= 47):
        return {
            'DB.TN.LXX.2B1.01': 4,
            'DB.TN.LXX.1B1.01': 4,
            'DB.TN.LXX.3B1.01': 1
        }
    
    return None

# Special ESMDBs
def get_esmdb_copy_counts(parent_name):
    """Get copy counts for special ESMDBs"""
    normalized = parent_name.replace('.01', '')
    
    if normalized == 'BB.05' or parent_name == 'BB.05' or normalized == 'EMDB.GF' or parent_name == 'EMDB.GF.01':
        return {
            'ESMDB.LL.RF.01(LIFT)': 1,
            'ESMDB.LL.RF.02(LIFT)': 1
        }
    
    return None

def normalize_mdb(mdb_value):
    """Normalize MDB values (MDB -> MDB1, MDB.GF.04 -> MDB4)"""
    if not mdb_value or not isinstance(mdb_value, str):
        return mdb_value or ''
    
    trimmed = mdb_value.strip().upper()
    if trimmed == 'MDB':
        return 'MDB1'
    if trimmed in ['MDB.GF.04', 'MDB GF 04']:
        return 'MDB4'
    return mdb_value

def parse_load_value(load_str):
    """Parse load value and extract numeric kW (exclude KVAR)"""
    if not load_str or not isinstance(load_str, str):
        return 0
    
    upper_str = load_str.upper()
    if 'KVAR' in upper_str:
        return 0
    
    import re
    match = re.search(r'(\d+\.?\d*)', load_str)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return 0
    return 0

def find_column_index(worksheet, search_text, header_row=1):
    """Find column index by searching header"""
    for col_idx, cell in enumerate(worksheet[header_row], 1):
        if cell.value:
            cell_value = str(cell.value).lower().strip()
            if search_text.lower() in cell_value:
                return col_idx
    return None

def generate_boq2(input_file, output_file):
    """Generate BOQ2 workbook with better tabularization"""
    print(f"Loading workbook: {input_file}")
    workbook = load_workbook(input_file, data_only=True)
    
    # Find TOTALLIST sheet
    total_list_sheet = None
    for sheet_name in workbook.sheetnames:
        if 'total' in sheet_name.lower() and 'list' in sheet_name.lower():
            total_list_sheet = workbook[sheet_name]
            break
    
    if not total_list_sheet:
        print("Error: TOTALLIST sheet not found")
        return False
    
    print(f"Found TOTALLIST sheet: {total_list_sheet.title}")
    
    # Read all data from TOTALLIST
    all_data = []
    headers = []
    
    # Find header row
    header_row_idx = 1
    for row_idx in range(1, min(10, total_list_sheet.max_row + 1)):
        row = total_list_sheet[row_idx]
        if any(cell.value for cell in row):
            headers = [str(cell.value).strip() if cell.value else '' for cell in row]
            header_row_idx = row_idx
            break
    
    # Find column indices
    itemdrop_col = find_column_index(total_list_sheet, 'itemdrop', header_row_idx) or 1
    mdb_col = find_column_index(total_list_sheet, 'mdb', header_row_idx)
    kind_col = find_column_index(total_list_sheet, 'kind', header_row_idx)
    fed_from_col = find_column_index(total_list_sheet, 'fed from', header_row_idx)
    load_col = find_column_index(total_list_sheet, 'load', header_row_idx)
    estimate_col = find_column_index(total_list_sheet, 'estimate', header_row_idx)
    no_of_items_col = find_column_index(total_list_sheet, 'no of items', header_row_idx)
    
    # Read data rows
    for row_idx in range(header_row_idx + 1, total_list_sheet.max_row + 1):
        row = total_list_sheet[row_idx]
        itemdrop = str(row[itemdrop_col - 1].value).strip() if row[itemdrop_col - 1].value else ''
        
        # Skip empty rows and total rows
        if not itemdrop or itemdrop.upper() in ['TOTAL', 'NET TOTAL', 'SUM AFTER']:
            continue
        
        item = {}
        item['Itemdrop'] = itemdrop
        item['MDB'] = normalize_mdb(str(row[mdb_col - 1].value).strip() if mdb_col and row[mdb_col - 1].value else '')
        item['KIND'] = str(row[kind_col - 1].value).strip() if kind_col and row[kind_col - 1].value else ''
        item['FED FROM'] = str(row[fed_from_col - 1].value).strip() if fed_from_col and row[fed_from_col - 1].value else ''
        item['Load'] = str(row[load_col - 1].value).strip() if load_col and row[load_col - 1].value else '0 kW'
        
        # Handle estimate with error checking
        if estimate_col and row[estimate_col - 1].value:
            try:
                item['Estimate'] = float(row[estimate_col - 1].value)
            except (ValueError, TypeError):
                item['Estimate'] = 0
        else:
            item['Estimate'] = 0
        
        # Handle NO OF ITEMS with error checking
        if no_of_items_col and row[no_of_items_col - 1].value:
            try:
                item['NO OF ITEMS'] = float(row[no_of_items_col - 1].value)
            except (ValueError, TypeError):
                item['NO OF ITEMS'] = 1
        else:
            item['NO OF ITEMS'] = 1
        
        all_data.append(item)
    
    print(f"Loaded {len(all_data)} items from TOTALLIST")
    
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ2"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    mdb_fill = PatternFill(start_color="C92A2A", end_color="C92A2A", fill_type="solid")
    mdb_font = Font(bold=True, color="FFFFFF", size=10)
    smdb_fill = PatternFill(start_color="2D8659", end_color="2D8659", fill_type="solid")
    smdb_font = Font(bold=True, color="FFFFFF", size=10)
    esmdb_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    db_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    special_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Level', 'Item Name', 'Kind', 'MDB', 'Fed From', 'Load (kW)', 'No. of Items', 'Estimate (AED)', 'Remarks']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 30
    
    current_row = 2
    
    # Special DBs
    special_dbs = ['DB.TN.LXX.1B1.01', 'DB.TN.LXX.2B1.01', 'DB.TN.LXX.3B1.01', 'DB.TH.GF.01']
    special_esmdbs = ['ESMDB.LL.RF.01(LIFT)', 'ESMDB.LL.RF.02(LIFT)']
    
    # Build hierarchical structure
    rmu_items = []
    mdb_items = {}
    smdb_items = {}
    esmdb_items = {}
    db_items = {}
    other_items = []
    
    # Organize items by hierarchy
    processed_item_names = set()
    
    for item in all_data:
        item_name = item['Itemdrop']
        kind = item['KIND'].upper() if item['KIND'] else ''
        mdb = item['MDB']
        fed_from = item['FED FROM']
        
        # Normalize item name for MDBs
        normalized_name = normalize_mdb(item_name)
        
        if item_name == 'RMU' or normalized_name == 'RMU':
            rmu_items.append(item)
            processed_item_names.add(item_name)
        elif kind == 'MDB':
            # Store with both original and normalized name
            mdb_items[item_name] = item
            if normalized_name != item_name:
                mdb_items[normalized_name] = item
            processed_item_names.add(item_name)
        elif kind == 'SMDB':
            smdb_items[item_name] = item
            processed_item_names.add(item_name)
        elif kind == 'ESMDB':
            esmdb_items[item_name] = item
            processed_item_names.add(item_name)
        elif kind == 'DB':
            db_items[item_name] = item
            processed_item_names.add(item_name)
        else:
            other_items.append(item)
            processed_item_names.add(item_name)
    
    # Helper function to write a row
    def write_row(level, item_name, kind, mdb, fed_from, load, no_of_items, estimate, remarks='', 
                  name_fill=None, name_font=None, kind_fill=None, kind_font=None):
        nonlocal current_row
        ws.cell(current_row, 1, str(level)).border = border
        cell = ws.cell(current_row, 2, item_name)
        if name_fill:
            cell.fill = name_fill
        if name_font:
            cell.font = name_font
        else:
            cell.font = Font(size=10)
        cell.border = border
        
        kind_cell = ws.cell(current_row, 3, kind)
        if kind_fill:
            kind_cell.fill = kind_fill
        if kind_font:
            kind_cell.font = kind_font
        kind_cell.border = border
        
        ws.cell(current_row, 4, mdb).border = border
        ws.cell(current_row, 5, fed_from).border = border
        ws.cell(current_row, 6, load).border = border
        ws.cell(current_row, 7, no_of_items).border = border
        ws.cell(current_row, 8, estimate).border = border
        ws.cell(current_row, 9, remarks).border = border
        current_row += 1
    
    # Write RMU
    if rmu_items:
        for item in rmu_items:
            write_row(
                '1', item['Itemdrop'], item['KIND'], item['MDB'], item['FED FROM'],
                item['Load'], item['NO OF ITEMS'], item['Estimate'],
                name_font=Font(bold=True, size=12),
                kind_fill=PatternFill(start_color="7D3C98", end_color="7D3C98", fill_type="solid"),
                kind_font=Font(bold=True, color="FFFFFF")
            )
    
    # Write MDBs
    mdb_order = ['MDB1', 'MDB2', 'MDB3', 'MDB4']
    for mdb_name in mdb_order:
        # Find MDB item (could be normalized name)
        mdb_item = None
        for key, item in mdb_items.items():
            normalized_key = normalize_mdb(key)
            if normalized_key == mdb_name or key == mdb_name:
                mdb_item = item
                break
        
        if mdb_item:
            write_row(
                '2', f"  {mdb_name}", mdb_item['KIND'], mdb_item['MDB'], mdb_item['FED FROM'],
                mdb_item['Load'], mdb_item['NO OF ITEMS'], mdb_item['Estimate'],
                name_font=Font(bold=True),
                kind_fill=mdb_fill,
                kind_font=mdb_font
            )
            
            # Write children of this MDB
            # Find SMDBs/ESMDBs fed from this MDB (check both normalized and original MDB names)
            for smdb_name, smdb_item in smdb_items.items():
                fed_from = smdb_item['FED FROM'] or ''
                # Check if this SMDB is fed from the current MDB (handle both MDB1/MDB and MDB4/MDB.GF.04)
                is_fed_from_mdb = False
                if mdb_name in fed_from:
                    is_fed_from_mdb = True
                elif mdb_name == 'MDB1' and ('MDB' in fed_from and 'MDB1' not in fed_from):
                    # Check if fed from just "MDB" (which should be MDB1)
                    fed_from_parts = [p.strip() for p in fed_from.split('\n')]
                    if 'MDB' in fed_from_parts and not any('MDB2' in p or 'MDB3' in p or 'MDB4' in p for p in fed_from_parts):
                        is_fed_from_mdb = True
                elif mdb_name == 'MDB4' and ('MDB.GF.04' in fed_from or 'MDB GF 04' in fed_from):
                    is_fed_from_mdb = True
                
                if is_fed_from_mdb:
                    # Write SMDB
                    write_row(
                        '3', f"    {smdb_name}", smdb_item['KIND'], smdb_item['MDB'], smdb_item['FED FROM'],
                        smdb_item['Load'], smdb_item['NO OF ITEMS'], smdb_item['Estimate'],
                        name_font=Font(bold=True),
                        kind_fill=smdb_fill,
                        kind_font=smdb_font
                    )
                    
                    # Check for special DBs under this SMDB
                    copy_counts = get_db_copy_counts(smdb_name)
                    if copy_counts:
                        for db_name, count in copy_counts.items():
                            if db_name in db_items:
                                db_item = db_items[db_name]
                                write_row(
                                    '4', f"      {db_name} ({count} copies)", db_item['KIND'], db_item['MDB'], smdb_name,
                                    db_item['Load'], count, db_item['Estimate'],
                                    f"Special DB - {count} copies under {smdb_name}",
                                    name_fill=special_fill,
                                    kind_fill=db_fill
                                )
                    
                    # Write regular DBs under this SMDB
                    for db_name, db_item in db_items.items():
                        if db_name not in special_dbs:
                            db_fed_from = db_item['FED FROM'] or ''
                            # Check if DB is fed from this SMDB (could be in multi-line FED FROM)
                            if smdb_name in db_fed_from:
                                write_row(
                                    '4', f"      {db_name}", db_item['KIND'], db_item['MDB'], db_item['FED FROM'],
                                    db_item['Load'], db_item['NO OF ITEMS'], db_item['Estimate'],
                                    kind_fill=db_fill
                                )
                    
                    # Write ESMDBs fed from this SMDB
                    for esmdb_name, esmdb_item in esmdb_items.items():
                        if smdb_name in esmdb_item['FED FROM']:
                            write_row(
                                '4', f"      {esmdb_name}", esmdb_item['KIND'], esmdb_item['MDB'], esmdb_item['FED FROM'],
                                esmdb_item['Load'], esmdb_item['NO OF ITEMS'], esmdb_item['Estimate'],
                                kind_fill=esmdb_fill
                            )
            
            # Write ESMDBs directly fed from MDB
            for esmdb_name, esmdb_item in esmdb_items.items():
                if mdb_name in esmdb_item['FED FROM'] and not any(smdb in esmdb_item['FED FROM'] for smdb in smdb_items.keys()):
                    write_row(
                        '3', f"    {esmdb_name}", esmdb_item['KIND'], esmdb_item['MDB'], esmdb_item['FED FROM'],
                        esmdb_item['Load'], esmdb_item['NO OF ITEMS'], esmdb_item['Estimate'],
                        kind_fill=esmdb_fill
                    )
            
            # Check for special DBs directly under MDB4 (DB.TH.GF.01)
            if mdb_name == 'MDB4':
                copy_counts = get_db_copy_counts('MDB4')
                if copy_counts:
                    for db_name, count in copy_counts.items():
                        if db_name in db_items:
                            db_item = db_items[db_name]
                            db_fed_from = db_item['FED FROM'] or ''
                            # Only show if MDB4 is in FED FROM (not already shown under SMDB)
                            if 'MDB4' in db_fed_from or 'MDB.GF.04' in db_fed_from or 'MDB GF 04' in db_fed_from:
                                write_row(
                                    '3', f"    {db_name} ({count} copies)", db_item['KIND'], db_item['MDB'], 'MDB4',
                                    db_item['Load'], count, db_item['Estimate'],
                                    f"Special DB - {count} copies under MDB4",
                                    name_fill=special_fill,
                                    kind_fill=db_fill
                                )
            
            # Write other items directly fed from MDB
            for other_item in other_items:
                if mdb_name in other_item['FED FROM']:
                    write_row(
                        '3', f"    {other_item['Itemdrop']}", other_item['KIND'], other_item['MDB'], other_item['FED FROM'],
                        other_item['Load'], other_item['NO OF ITEMS'], other_item['Estimate']
                    )
    
    # Handle special ESMDBs under BB.05 and EMDB.GF.01
    esmdb_parents = ['BB.05', 'EMDB.GF.01']
    for parent_name in esmdb_parents:
        copy_counts = get_esmdb_copy_counts(parent_name)
        if copy_counts:
            # Find parent item
            parent_item = None
            for item in all_data:
                if item['Itemdrop'] == parent_name:
                    parent_item = item
                    break
            
            if parent_item:
                # Check if parent is already written (might be under an MDB)
                # Write parent if not already written
                write_row(
                    '3', f"    {parent_name}", parent_item['KIND'], parent_item['MDB'], parent_item['FED FROM'],
                    parent_item['Load'], parent_item['NO OF ITEMS'], parent_item['Estimate'],
                    name_font=Font(bold=True),
                    kind_fill=PatternFill(start_color="98D8C8", end_color="98D8C8", fill_type="solid"),
                    kind_font=Font(bold=True)
                )
                
                # Write special ESMDBs
                for esmdb_name, count in copy_counts.items():
                    esmdb_item = esmdb_items.get(esmdb_name)
                    if esmdb_item:
                        write_row(
                            '4', f"      {esmdb_name} ({count} copies)", esmdb_item['KIND'], esmdb_item['MDB'], parent_name,
                            esmdb_item['Load'], count, esmdb_item['Estimate'],
                            f"Special ESMDB - {count} copies under {parent_name}",
                            name_fill=special_fill,
                            kind_fill=esmdb_fill
                        )
    
    # Write any remaining items that weren't processed
    remaining_items = [item for item in all_data if item['Itemdrop'] not in processed_item_names]
    if remaining_items:
        write_row('', '', '', '', '', '', '', '', '=== Remaining Items ===', name_font=Font(bold=True, italic=True))
        for item in remaining_items:
            write_row(
                '', item['Itemdrop'], item['KIND'], item['MDB'], item['FED FROM'],
                item['Load'], item['NO OF ITEMS'], item['Estimate']
            )
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Save workbook
    print(f"Saving BOQ2 workbook: {output_file}")
    wb.save(output_file)
    print(f"✓ BOQ2 workbook generated successfully!")
    return True

if __name__ == '__main__':
    input_file = 'e2.xlsx'
    output_file = 'BOQ2.xlsx'
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found")
        sys.exit(1)
    
    success = generate_boq2(input_file, output_file)
    if not success:
        sys.exit(1)

