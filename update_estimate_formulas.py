#!/usr/bin/env python3
"""
Script to update ESTIMATE column in TOTALLIST sheet with formulas
that reference NET TOTAL (F163) from respective sheets
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os
import re

def normalize_sheet_name(item_name):
    """Normalize item name to match sheet name (MDB -> MDB1, MDB.GF.04 -> MDB4)"""
    if not item_name or not isinstance(item_name, str):
        return item_name or ''
    
    trimmed = item_name.strip().upper()
    if trimmed == 'MDB':
        return 'MDB1'
    if trimmed in ['MDB.GF.04', 'MDB GF 04']:
        return 'MDB4'
    return item_name

def find_net_total_cell(worksheet):
    """Find the cell address where NET TOTAL is located (defaults to F163)"""
    # User specified F163 as the standard location
    # But we'll still try to find it dynamically for accuracy
    
    # Search for ITEM and AMOUNT columns
    item_col = None
    amount_col = None
    
    # Search header rows
    for row_idx in range(1, min(10, worksheet.max_row + 1)):
        for col_idx, cell in enumerate(worksheet[row_idx], 1):
            if cell.value:
                cell_value = str(cell.value).strip().upper()
                if 'ITEM' in cell_value or 'DESCRIPTION' in cell_value:
                    if item_col is None:
                        item_col = col_idx
                if 'AMOUNT' in cell_value and 'PRICE' not in cell_value:
                    if amount_col is None:
                        amount_col = col_idx
        
        if item_col and amount_col:
            break
    
    # Fallback to defaults
    if item_col is None:
        item_col = 3  # Column C
    if amount_col is None:
        amount_col = 6  # Column F (F column)
    
    # Check if F163 exists and has NET TOTAL
    if worksheet.max_row >= 163:
        item_cell = worksheet.cell(163, item_col)
        if item_cell.value:
            item_value = str(item_cell.value).strip().upper()
            if 'NET TOTAL' in item_value:
                cell_address = f"{get_column_letter(amount_col)}163"
                return cell_address
    
    # Search for NET TOTAL from bottom up (around row 163 area)
    search_start = min(worksheet.max_row, 200)
    search_end = max(1, search_start - 50)
    
    for row_idx in range(search_start, search_end, -1):
        item_cell = worksheet.cell(row_idx, item_col)
        if item_cell.value:
            item_value = str(item_cell.value).strip().upper()
            if 'NET TOTAL' in item_value:
                # Return the cell address of the AMOUNT column for this row
                cell_address = f"{get_column_letter(amount_col)}{row_idx}"
                return cell_address
    
    # Default to F163 as specified by user
    return 'F163'

def find_sheet_name(item_name, workbook):
    """Find the sheet name that corresponds to an item"""
    normalized = normalize_sheet_name(item_name)
    sheet_names = workbook.sheetnames
    
    # Try exact match first
    if normalized in sheet_names:
        return normalized
    if item_name in sheet_names:
        return item_name
    
    # Try partial match
    for sheet_name in sheet_names:
        if sheet_name.upper() == normalized.upper():
            return sheet_name
        if normalized.upper() in sheet_name.upper() or sheet_name.upper() in normalized.upper():
            return sheet_name
    
    # Try base name matching
    base_name = normalized.split('.')[0] if '.' in normalized else normalized
    for sheet_name in sheet_names:
        if base_name.upper() in sheet_name.upper():
            return sheet_name
    
    return None

def update_estimate_formulas(file_path):
    """Update ESTIMATE column with formulas referencing respective sheets"""
    try:
        print(f"Loading workbook: {file_path}")
        workbook = load_workbook(file_path, data_only=False)  # Keep formulas
        
        # Find TOTALLIST sheet
        total_list_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'total' in sheet_name.lower() and 'list' in sheet_name.lower():
                total_list_sheet = workbook[sheet_name]
                break
        
        if not total_list_sheet:
            print("Error: TOTALLIST sheet not found")
            return False
        
        print(f"Found TOTALLIST sheet: {total_list_sheet.title}")
        
        # Find columns
        header_row = 1
        itemdrop_col = None
        estimate_col = None
        
        # Find header row and columns
        for row_idx in range(1, min(10, total_list_sheet.max_row + 1)):
            for col_idx, cell in enumerate(total_list_sheet[row_idx], 1):
                if cell.value:
                    cell_value = str(cell.value).lower().strip()
                    if 'itemdrop' in cell_value:
                        itemdrop_col = col_idx
                    if 'estimate' in cell_value:
                        estimate_col = col_idx
            
            if itemdrop_col and estimate_col:
                header_row = row_idx
                break
        
        if not itemdrop_col:
            itemdrop_col = 1
        if not estimate_col:
            print("Error: ESTIMATE column not found")
            return False
        
        print(f"Found columns - Itemdrop: {get_column_letter(itemdrop_col)}, Estimate: {get_column_letter(estimate_col)}")
        
        updated_count = 0
        skipped_count = 0
        not_found_count = 0
        
        # Process each row
        for row_idx in range(header_row + 1, total_list_sheet.max_row + 1):
            itemdrop_cell = total_list_sheet.cell(row_idx, itemdrop_col)
            estimate_cell = total_list_sheet.cell(row_idx, estimate_col)
            
            if not itemdrop_cell.value:
                continue
            
            item_name = str(itemdrop_cell.value).strip()
            
            # Skip empty rows and total rows
            if not item_name or item_name.upper() in ['TOTAL', 'NET TOTAL', 'SUM AFTER']:
                continue
            
            # Skip RMU row
            if item_name.upper() == 'RMU':
                skipped_count += 1
                continue
            
            # Find corresponding sheet
            sheet_name = find_sheet_name(item_name, workbook)
            
            if not sheet_name:
                print(f"  ⚠ Sheet not found for: {item_name}")
                not_found_count += 1
                continue
            
            # Get the worksheet to find NET TOTAL cell
            try:
                worksheet = workbook[sheet_name]
                net_total_cell = find_net_total_cell(worksheet)
                
                # Create formula: ='SheetName'!F163
                # Escape sheet name if it contains special characters
                safe_sheet_name = sheet_name
                if ' ' in sheet_name or '-' in sheet_name or any(c in sheet_name for c in ['(', ')', '[', ']', '.']):
                    safe_sheet_name = f"'{sheet_name}'"
                
                formula = f"={safe_sheet_name}!{net_total_cell}"
                estimate_cell.value = formula
                
                # Only print if not F163 (to reduce output noise)
                if net_total_cell != 'F163':
                    print(f"  ✓ Row {row_idx}: {item_name} -> ={safe_sheet_name}!{net_total_cell}")
                updated_count += 1
                
            except Exception as e:
                print(f"  ✗ Error processing {item_name} -> {sheet_name}: {e}")
                not_found_count += 1
        
        # Save workbook
        print(f"\nSaving workbook...")
        workbook.save(file_path)
        
        print(f"\n✓ Update complete!")
        print(f"  Updated: {updated_count} rows")
        print(f"  Skipped: {skipped_count} rows")
        print(f"  Not found: {not_found_count} rows")
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    file_path = 'e2.xlsx'
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found")
        sys.exit(1)
    
    success = update_estimate_formulas(file_path)
    if not success:
        sys.exit(1)

