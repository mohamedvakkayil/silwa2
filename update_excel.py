#!/usr/bin/env python3
"""
Script to copy net total values from corresponding sheets 
to the estimate column in the total list sheet
"""

import openpyxl
from openpyxl import load_workbook
import sys

def find_net_total(worksheet):
    """Find the net total value in a worksheet"""
    net_total = None
    
    # First, find the ITEM and AMOUNT column indices by searching headers
    item_col = None
    amount_col = None
    
    # Search header rows (usually row 1-3)
    for row_idx in range(1, min(5, worksheet.max_row + 1)):
        for col_idx, cell in enumerate(worksheet[row_idx], 1):
            if cell.value:
                cell_value = str(cell.value).strip().upper()
                if 'ITEM' in cell_value or 'DESCRIPTION' in cell_value:
                    if item_col is None:
                        item_col = col_idx
                if 'AMOUNT' in cell_value and 'PRICE' not in cell_value:
                    if amount_col is None:
                        amount_col = col_idx
    
    # Fallback to known positions if not found
    if item_col is None:
        item_col = 3  # Default ITEM column
    if amount_col is None:
        amount_col = 6  # Default AMOUNT column
    
    # Search from bottom up (totals are usually at the end)
    for row_idx in range(worksheet.max_row, max(1, worksheet.max_row - 50), -1):
        item_cell = worksheet.cell(row_idx, item_col)
        if item_cell.value:
            item_value = str(item_cell.value).strip().upper()
            if 'NET TOTAL' in item_value:
                # Get value from AMOUNT column
                amount_cell = worksheet.cell(row_idx, amount_col)
                if amount_cell.value is not None:
                    try:
                        net_total = float(amount_cell.value)
                        print(f"  Found NET TOTAL at row {row_idx}, column {amount_col}: {net_total}")
                        break
                    except (ValueError, TypeError):
                        pass
    
    return net_total

def find_estimate_column(worksheet):
    """Find the estimate column index"""
    # Search header row (usually row 1)
    for row_idx in range(1, min(5, worksheet.max_row + 1)):
        for col_idx, cell in enumerate(worksheet[row_idx], 1):
            if cell.value:
                cell_value = str(cell.value).lower().strip()
                if 'estimate' in cell_value:
                    return col_idx
    
    return None

def find_name_column(worksheet):
    """Find the column that contains item names (Itemdrop)"""
    # Search header row
    for row_idx in range(1, min(5, worksheet.max_row + 1)):
        for col_idx, cell in enumerate(worksheet[row_idx], 1):
            if cell.value:
                cell_value = str(cell.value).lower().strip()
                # Prioritize Itemdrop, then other name fields
                if 'itemdrop' in cell_value:
                    return col_idx
                elif 'item' in cell_value and 'drop' not in cell_value:
                    return col_idx
                elif 'name' in cell_value or 'description' in cell_value:
                    return col_idx
    
    # Default to first column
    return 1

def update_excel(file_path):
    """Main function to update Excel file"""
    try:
        print(f"Loading workbook: {file_path}")
        workbook = load_workbook(file_path)
        
        # Find total list sheet
        total_list_sheet = None
        sheet_names = workbook.sheetnames
        
        print(f"Available sheets: {', '.join(sheet_names)}")
        
        # Look for "total list" sheet
        for sheet_name in sheet_names:
            if 'total' in sheet_name.lower() and 'list' in sheet_name.lower():
                total_list_sheet = workbook[sheet_name]
                print(f"Found total list sheet: {sheet_name}")
                break
        
        if total_list_sheet is None:
            # Try to find it by other names
            for sheet_name in sheet_names:
                if 'list' in sheet_name.lower():
                    total_list_sheet = workbook[sheet_name]
                    print(f"Using sheet as total list: {sheet_name}")
                    break
        
        if total_list_sheet is None:
            print("ERROR: Could not find total list sheet!")
            return False
        
        # Find estimate column
        estimate_col = find_estimate_column(total_list_sheet)
        if estimate_col is None:
            print("ERROR: Could not find estimate column!")
            return False
        
        print(f"Found estimate column at column {estimate_col}")
        
        # Find name column
        name_col = find_name_column(total_list_sheet)
        print(f"Found name column at column {name_col}")
        
        # Create mapping of sheet names to net totals
        sheet_net_totals = {}
        
        for sheet_name in sheet_names:
            if sheet_name == total_list_sheet.title:
                continue
            
            worksheet = workbook[sheet_name]
            net_total = find_net_total(worksheet)
            
            if net_total is not None:
                sheet_net_totals[sheet_name] = net_total
                print(f"Sheet '{sheet_name}': Net Total = {net_total}")
            else:
                print(f"Sheet '{sheet_name}': Could not find net total")
        
        # Update estimate column in total list sheet
        updated_count = 0
        skipped_count = 0
        not_found_count = 0
        
        # Start from row 2 (assuming row 1 is header)
        for row_idx in range(2, total_list_sheet.max_row + 1):
            name_cell = total_list_sheet.cell(row_idx, name_col)
            estimate_cell = total_list_sheet.cell(row_idx, estimate_col)
            
            if not name_cell.value:
                continue
            
            item_name = str(name_cell.value).strip()
            
            # Skip empty rows, totals, and special rows
            if not item_name or item_name.upper() in ['TOTAL', 'NET TOTAL', 'SUM', '']:
                skipped_count += 1
                continue
            
            # Normalize item name for matching (MDB -> MDB1, MDB.GF.04 -> MDB4)
            normalized_item_name = item_name
            if item_name.upper() == 'MDB':
                normalized_item_name = 'MDB1'
            elif item_name.upper() in ['MDB.GF.04', 'MDB GF 04']:
                normalized_item_name = 'MDB4'
            
            # Try to match with sheet names
            matched = False
            best_match = None
            best_match_score = 0
            
            for sheet_name, net_total in sheet_net_totals.items():
                # Normalize sheet name for matching
                normalized_sheet_name = sheet_name
                if sheet_name.upper() == 'MDB':
                    normalized_sheet_name = 'MDB1'
                elif sheet_name.upper() in ['MDB.GF.04', 'MDB GF 04']:
                    normalized_sheet_name = 'MDB4'
                
                # Exact match (highest priority) - try both original and normalized
                if (sheet_name.lower() == item_name.lower() or 
                    normalized_sheet_name.lower() == normalized_item_name.lower() or
                    sheet_name.lower() == normalized_item_name.lower() or
                    normalized_sheet_name.lower() == item_name.lower()):
                    estimate_cell.value = net_total
                    updated_count += 1
                    matched = True
                    print(f"✓ Row {row_idx}: {item_name} = {net_total} (exact match with {sheet_name})")
                    break
                
                # Calculate match score for partial matches
                item_lower = item_name.lower()
                sheet_lower = sheet_name.lower()
                
                # Check if item name is contained in sheet name or vice versa
                if item_lower in sheet_lower:
                    score = len(item_lower) / len(sheet_lower)
                    if score > best_match_score:
                        best_match_score = score
                        best_match = (sheet_name, net_total)
                elif sheet_lower in item_lower:
                    score = len(sheet_lower) / len(item_lower)
                    if score > best_match_score:
                        best_match_score = score
                        best_match = (sheet_name, net_total)
            
            # If no exact match, try best partial match
            if not matched and best_match and best_match_score > 0.5:
                sheet_name, net_total = best_match
                estimate_cell.value = net_total
                updated_count += 1
                print(f"✓ Row {row_idx}: {item_name} = {net_total} (partial match: {best_match_score:.2f} with {sheet_name})")
                matched = True
            
            # Try matching by base name (before first dot) if still not matched
            if not matched:
                base_name = item_name.split('.')[0] if '.' in item_name else item_name
                for sheet_name, net_total in sheet_net_totals.items():
                    sheet_base = sheet_name.split('.')[0] if '.' in sheet_name else sheet_name
                    if base_name.lower() == sheet_base.lower():
                        estimate_cell.value = net_total
                        updated_count += 1
                        matched = True
                        print(f"✓ Row {row_idx}: {item_name} = {net_total} (base name match with {sheet_name})")
                        break
            
            if not matched:
                not_found_count += 1
                print(f"✗ Row {row_idx}: {item_name} - No matching sheet found")
        
        # Save the workbook
        print(f"\n{'='*60}")
        print(f"Summary:")
        print(f"  Updated: {updated_count} rows")
        print(f"  Skipped: {skipped_count} rows (totals/empty)")
        print(f"  Not found: {not_found_count} rows")
        print(f"{'='*60}")
        print(f"\nSaving workbook...")
        workbook.save(file_path)
        print("✓ Done! Excel file updated successfully.")
        return True
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "e2.xlsx"
    update_excel(file_path)

